import eumdac
import requests
import datetime
import shutil
import time
import os
import xarray as xr
import matplotlib.pyplot as plt
import cartopy
import xarray as xr
import numpy as np
import pandas as pd
import os
import urllib3
from pathlib import Path
import pandas as pd
import zipfile
import gzip
import glob
import satpy
from satpy.scene import Scene
from satpy import find_files_and_readers
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
import xarray as xr
import rioxarray
import numpy as np
from pyproj import Geod

def GetValues(Filename, time, lat, lon, radius_km=10):
    ds = xr.open_dataset("./" + Filename)
    
    da = ds["ir_38"]
    db = ds["ir_105"]
    
    t0 = np.datetime64(time)
    
    # Requested time window ±30 min
    time_window = slice(t0 - np.timedelta64(30, "m"), t0 + np.timedelta64(30, "m"))
    
    # Select data in that window
    da_sel = da.sel(time=time_window)
    db_sel = db.sel(time=time_window)
    
    # If the slice is empty, fall back to the entire available time
    if da_sel.time.size == 0:
        print(f"No data in ±30 min window for {Filename} at {time}, using available time instead")
        da_sel = da
        db_sel = db
    
    # Min over time
    da_time_min = da_sel.min(dim="time")
    db_time_min = db_sel.min(dim="time")
    
    # 2D lat/lon arrays
    lats = da.latitude.values
    lons = da.longitude.values
    
    # Convert to radians
    lat_rad = np.radians(lat)
    lon_rad = np.radians(lon)
    lats_rad = np.radians(lats)
    lons_rad = np.radians(lons)
    
    # Haversine distance
    dlat = lats_rad - lat_rad
    dlon = lons_rad - lon_rad
    a = np.sin(dlat/2)**2 + np.cos(lat_rad) * np.cos(lats_rad) * np.sin(dlon/2)**2
    c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1-a))
    R = 6371
    dist_matrix = R * c
    
    # Mask within radius
    mask = dist_matrix <= radius_km
    
    # Apply mask
    da_time_masked = da_time_min.where(mask)
    db_time_masked = db_time_min.where(mask)
    
    amin_value = da_time_masked.min().item()
    bmin_value = db_time_masked.min().item()
    
    print(amin_value, bmin_value)
    return amin_value, bmin_value

wb = load_workbook("Fire_spread_events_final_review_MSG.xlsx")
ws = wb.active

ws["DB1"] = "MTG_BT_38_MIN"
ws["DC1"] = "MTG_BT_105_MIN"


# Fill values row by row
for row in range(2, ws.max_row + 1):
    Case=ws[f"C{row}"].value
    Filename=Case+str(".nc")
    if os.path.isfile("./"+Filename):
        ## Get values close to each time step
        
        #Get time step
        time = Case=ws[f"V{row}"].value
        print(time)
        
        #Get point
        lat = float(ws[f"Q{row}"].value)
        lon = float(ws[f"R{row}"].value)
        
        #Get value
        print(f"{Filename} timestep:{time} lat:{lat} lon:{lon}")
        bt39min,bt108min = GetValues(Filename,time,lat,lon)
        ws[f"DB{row}"] = bt39min
        ws[f"DC{row}"] = bt108min
    # ws[f"C{row}"] = 100
    # ws[f"D{row}"] = ws[f"A{row}"].value * 2

wb.save("Fire_spread_events_final_review_MSG_MTG.xlsx")
